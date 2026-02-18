"""
総合ライフダッシュボード生成スクリプト
Obsidian日記から睡眠・筋トレ・歩数・読書データを抽出し、
HTMLダッシュボード + Obsidianレポートを生成する

使い方:
  python life_dashboard.py            # 生成のみ
  python life_dashboard.py --deploy   # 生成 + GitHub Pagesにデプロイ
"""

import re
import json
import sys
import os
import subprocess
import argparse
from pathlib import Path
from datetime import datetime, timedelta

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# === 設定 ===
DIARY_DIR = Path(r"C:\Documents\Obsidian Vault\Main Vault\日記")
VAULT_DIR = Path(r"C:\Documents\Obsidian Vault\Main Vault")
SCRIPT_DIR = Path(__file__).parent
DOCS_DIR = SCRIPT_DIR / "docs"

# 全角数字→半角
ZEN_TO_HAN = str.maketrans('０１２３４５６７８９', '0123456789')

def zen_to_int(s: str) -> int:
    """全角数字を含む文字列から整数を取得"""
    s = s.translate(ZEN_TO_HAN)
    m = re.search(r'(\d+)', s)
    return int(m.group(1)) if m else 0


def parse_frontmatter(text: str) -> dict:
    fm = {}
    if not text.startswith("---"):
        return fm
    end = text.find("---", 3)
    if end == -1:
        return fm
    block = text[3:end].strip()
    for line in block.splitlines():
        line = line.strip()
        if line.startswith("-") or not line:
            continue
        if ":" in line:
            key, _, val = line.partition(":")
            fm[key.strip()] = val.strip().strip('"').strip("'")
    return fm


def parse_duration(s: str) -> float:
    hours, mins = 0, 0
    hm = re.search(r'(\d+)\s*時間', s)
    if hm: hours = int(hm.group(1))
    mm = re.search(r'(\d+)\s*分', s)
    if mm: mins = int(mm.group(1))
    return hours + mins / 60


def parse_sleep_details(text: str) -> dict:
    d = {}
    m = re.search(r'睡眠スコア[：:]?\s*(\d+)', text)
    if m: d['score'] = int(m.group(1))

    m = re.search(r'深い\s*([\d時間分]+)', text)
    if m: d['deep'] = round(parse_duration(m.group(1)), 2)
    m = re.search(r'ライト\s*([\d時間分]+)', text)
    if m: d['light'] = round(parse_duration(m.group(1)), 2)
    m = re.search(r'レム\s*([\d時間分]+)', text)
    if m: d['rem'] = round(parse_duration(m.group(1)), 2)
    m = re.search(r'覚醒\s*([\d時間分]+)', text)
    if m: d['awake'] = round(parse_duration(m.group(1)), 2)

    m = re.search(r'就寝\s*(\d{1,2}:\d{2})\s*[〜~～]\s*起床\s*(\d{1,2}:\d{2})', text)
    if m:
        d['bedtime'] = m.group(1)
        d['waketime'] = m.group(2)
    else:
        m = re.search(r'(\d{1,2}:\d{2})\s*[〜~～]\s*(\d{1,2}:\d{2})', text)
        if m:
            d['bedtime'] = m.group(1)
            d['waketime'] = m.group(2)

    m = re.search(r'天気::(.+)', text)
    if m:
        weather = re.split(r'[。、\.!！]', m.group(1).strip())[0].strip()
        d['weather'] = weather

    m = re.search(r'気分::(.+)', text, re.DOTALL)
    if m:
        # Get full mood text (multi-line until next section)
        mood_raw = m.group(1).strip()
        # Cut at next section header
        cut = re.search(r'\n-\s+\S+::', mood_raw)
        if cut:
            mood_raw = mood_raw[:cut.start()].strip()
        d['mood'] = mood_raw

    m = re.search(r'歩数::\s*([\d,]+)\s*歩', text)
    if m: d['steps'] = int(m.group(1).replace(',', ''))

    return d


def parse_exercise(fm: dict) -> dict:
    ex = {}
    for key, field in [('スクワット', 'squat'), ('腹筋', 'abs'), ('腕立て伏せ', 'pushup')]:
        val = fm.get(key, '')
        if val:
            ex[field] = zen_to_int(val)
    return ex


def parse_reading(text: str) -> list:
    books = []
    in_reading = False
    for line in text.splitlines():
        if '今日読んだ本' in line or '📚' in line:
            in_reading = True
            continue
        if in_reading:
            if line.strip().startswith('#') or (line.strip() and not line.strip().startswith('-')):
                break
            m = re.search(r'\[\[(.+?)(?:\|.+?)?\]\]', line)
            if m:
                title = m.group(1)
                finished = '読了' in line
                books.append({'title': title, 'finished': finished})
    return books


# === ジャンル分類 ===

GENRE_RULES = [
    # (ジャンル名, 著者キーワード, タイトルキーワード)
    ('ミステリー', ['道尾秀介','東野圭吾','今村昌弘','阿津川辰海','我孫子武丸','浦賀和宏',
                    '詠坂雄二','似鳥鶏','東川篤哉','西式豊','西式 豊','西澤保彦','五十嵐律人',
                    '潮谷験','大山誠一郎','知念実希人','三日市零','村上暢','早坂吝',
                    '紺野天龍','神永学','誉田哲也','住田祐','小倉千明','田村和大',
                    'クレイヴン','フリーダ','梨'], 
                   ['殺人','ミステリー','密室','探偵','八雲','Jミステリー','カラスの親指',
                    'カエルの小指','マスカレード','ロンド','Another','ＡＮＯＴＨＥＲＳの殺人',
                    'シンデレラ城','ハウスメイド','操る男','亡霊','仕掛島','ラットマン',
                    '復讐は','硝子の塔','白鷺立つ','デスチェア','嘘つき','兇人邸',
                    '透明人間は密室','不在の生存証明','迷宮牢','推理大戦','裁く眼',
                    'にいたる病','身から出た闇','時空犯','幻告']),
    ('仏教・宗教', ['梶山雄一','四夷法顕','菊地章太','玄侑宗久','平雅行'],
                   ['仏教','浄土','輪廻','華厳','鎌倉仏教','儒教','道教','涅槃','衆生']),
    ('ホラー・怪奇', ['背筋','小松左京','小松 左京'],
                     ['恐怖','ＳＦ','牛の首','ホラー','心霊']),
    ('自己啓発・学習', ['樺沢紫苑','榎本博明','石田光規','徳谷智史','井上慎平','西岡壱誠',
                       'ベンジャミン','山口 周','山口周','サルマン・カーン','八木','キム・イッカン',
                       'アダム・グラント','大塚あみ','出口治明','毛内拡'],
                      ['勉強','集中力','全力化','自己成長','経営','HIDDEN','読書を仕事',
                       '読書する脳','100日チャレンジ','世界一ゆるい','巨人のノート',
                       'セカンド・チャンス','可能性の科学']),
    ('健康・科学', ['池田光史','石川泰弘','東島威史','稲葉俊郎','リーバーマン'],
                   ['歩く','睡眠','不夜脳','運動の科学','メディスン','ぐっすり眠れる']),
    ('社会・ノンフィクション', ['パオロ','富永京子','宮下英樹','大城道則','オードリー','李雅卿'],
                              ['社会','ヘンなの','なぜ社会','古代文字','歴史','落とし穴',
                               '宅建士']),
]

def classify_genre(title: str) -> str:
    """タイトルからジャンルを推定"""
    for genre, authors, keywords in GENRE_RULES:
        for kw in keywords:
            if kw in title:
                return genre
        for author in authors:
            if author in title:
                return genre
    # テーマ性のある小説
    if re.search(r'[小説|物語|文庫]', title):
        return 'その他小説'
    return 'その他'


def get_returned_titles() -> set:
    """エクセルから返却済み＆未読了の本タイトルを取得"""
    excel_path = Path(r"C:\Users\trexa\OneDrive\記録\図書館にて借りた本の記録.xlsx")
    if not excel_path.exists():
        return set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        returned_not_finished = set()
        for name in wb.sheetnames:
            ws = wb[name]
            for row in range(2, ws.max_row + 1):
                title = ws.cell(row, 3).value
                returned = ws.cell(row, 8).value
                finished = ws.cell(row, 9).value
                if title and returned and str(returned).strip() == '\u2714':
                    if not (finished and str(finished).strip() == '\u2714'):
                        returned_not_finished.add(title.strip())
        return returned_not_finished
    except Exception as e:
        print(f"   ⚠️ エクセル読み込みエラー: {e}")
        return set()


def build_reading_summary(data: list[dict]) -> dict:
    """全日記から読書サマリーを構築（ジャンル別・ペース分析）"""
    book_tracker = {}  # title -> {first, last, finished, days_seen, genre}
    
    for entry in data:
        date_str = entry['date']
        for b in entry.get('books', []):
            title = b['title']
            if title not in book_tracker:
                book_tracker[title] = {
                    'title': title,
                    'first': date_str,
                    'last': date_str,
                    'finished': None,
                    'days_seen': 0,
                    'genre': classify_genre(title),
                }
            book_tracker[title]['last'] = date_str
            book_tracker[title]['days_seen'] += 1
            if b.get('finished'):
                book_tracker[title]['finished'] = date_str
    
    # 返却済み＆未読了の本を除外
    returned = get_returned_titles()
    # ユーザー確認済み返却本（エクセルにない分）
    manual_returned = {'オードリー・タンの母', '新アジア仏教史', '不可触民と現代インド',
                       '列島創世記', '日本の歴史1', '日本の歴史2', '日本史を宗教で読みなおす',
                       '今日の学び', '📒読書ノート', '買う食料・日用品', '欲しいもの', 'すること',
                       'セカンド・チャンス', '異次元緩和の罪と罰', '初めてのマルクス'}
    for b in list(book_tracker.values()):
        if not b['finished']:
            title_short = b['title'].split(' - ')[0]
            # Manual exclusion check
            for mr in manual_returned:
                if mr in title_short or title_short in mr:
                    b['returned'] = True
                    break
            if b.get('returned'):
                continue
            # Excel return check (improved matching)
            for rt in returned:
                if (title_short[:6] in rt or rt[:6] in title_short
                        or title_short in rt or rt in title_short):
                    b['returned'] = True
                    break
    
    # ペース計算
    all_books = [b for b in book_tracker.values() if not b.get('returned')]
    for b in all_books:
        if b['finished']:
            d1 = datetime.strptime(b['first'], '%Y-%m-%d')
            d2 = datetime.strptime(b['finished'], '%Y-%m-%d')
            b['reading_days'] = (d2 - d1).days + 1
        else:
            b['reading_days'] = None
    
    # ジャンル別集計
    genre_counts = {}
    genre_finished = {}
    for b in all_books:
        g = b['genre']
        genre_counts[g] = genre_counts.get(g, 0) + 1
        if b['finished']:
            genre_finished[g] = genre_finished.get(g, 0) + 1
    
    # 月別ジャンル読了
    monthly_genre = {}
    for b in all_books:
        if b['finished']:
            m = b['finished'][:7]
            if m not in monthly_genre:
                monthly_genre[m] = {}
            g = b['genre']
            monthly_genre[m][g] = monthly_genre[m].get(g, 0) + 1
    
    # ペース統計
    paces = [b['reading_days'] for b in all_books if b['reading_days'] is not None]
    avg_pace = sum(paces) / len(paces) if paces else 0
    
    returned_count = sum(1 for b in book_tracker.values() if b.get('returned'))
    print(f"   📕 返却済み（未読了）除外: {returned_count}冊")
    
    return {
        'books': all_books,
        'genre_counts': genre_counts,
        'genre_finished': genre_finished,
        'monthly_genre': monthly_genre,
        'avg_pace': round(avg_pace, 1),
        'total': len(all_books),
        'finished': sum(1 for b in all_books if b['finished']),
    }


def extract_all_data() -> list[dict]:
    entries = []
    for f in sorted(DIARY_DIR.glob("*.md")):
        m = re.match(r'(\d{4}-\d{2}-\d{2})', f.stem)
        if not m: continue
        date_str = m.group(1)
        text = f.read_text(encoding='utf-8')
        fm = parse_frontmatter(text)

        entry = {'date': date_str}
        
        # Sleep
        sleep_val = fm.get('sleep')
        if sleep_val:
            try:
                entry['hours'] = float(sleep_val)
            except ValueError:
                pass

        # Sleep details from body
        details = parse_sleep_details(text)
        entry.update(details)

        # Exercise from frontmatter
        exercise = parse_exercise(fm)
        if exercise:
            entry['exercise'] = exercise

        # Reading from body
        books = parse_reading(text)
        if books:
            entry['books'] = books

        # Only include if there's meaningful data
        if len(entry) > 1:
            entries.append(entry)

    return entries


# === 睡眠分析レポート ===

def generate_sleep_report(data: list[dict]) -> dict:
    today = datetime.now()
    
    def filter_range(start_date, end_date):
        return [d for d in data if d.get('hours') and start_date <= d['date'] <= end_date]

    this_week_start = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
    this_week_end = today.strftime('%Y-%m-%d')
    last_week_start = (today - timedelta(days=today.weekday() + 7)).strftime('%Y-%m-%d')
    last_week_end = (today - timedelta(days=today.weekday() + 1)).strftime('%Y-%m-%d')
    
    this_month = today.strftime('%Y-%m')
    last_month_dt = today.replace(day=1) - timedelta(days=1)
    last_month = last_month_dt.strftime('%Y-%m')

    this_week = filter_range(this_week_start, this_week_end)
    last_week = filter_range(last_week_start, last_week_end)
    this_month_data = [d for d in data if d.get('hours') and d['date'].startswith(this_month)]
    last_month_data = [d for d in data if d.get('hours') and d['date'].startswith(last_month)]

    def avg(arr): return sum(arr)/len(arr) if arr else 0
    def safe_avg(data_list, key):
        vals = [d[key] for d in data_list if key in d]
        return avg(vals) if vals else None

    report = {
        'generated': today.strftime('%Y-%m-%d %H:%M'),
        'weekly': {},
        'monthly_comparison': {},
        'improvements': [],
        'streaks': {},
    }

    # Weekly
    if this_week:
        wk = report['weekly']
        wk['avg_hours'] = round(avg([d['hours'] for d in this_week]), 2)
        wk['avg_score'] = round(safe_avg(this_week, 'score') or 0, 1)
        wk['days'] = len(this_week)
        wk['best_day'] = max(this_week, key=lambda d: d.get('score', 0)).get('date')
        wk['worst_day'] = min(this_week, key=lambda d: d.get('score', 100)).get('date')
        if last_week:
            lw_avg = avg([d['hours'] for d in last_week])
            wk['vs_last_week'] = round(wk['avg_hours'] - lw_avg, 2)

    # Monthly comparison
    if this_month_data and last_month_data:
        mc = report['monthly_comparison']
        mc['this_month'] = this_month
        mc['last_month'] = last_month
        mc['this_avg_hours'] = round(avg([d['hours'] for d in this_month_data]), 2)
        mc['last_avg_hours'] = round(avg([d['hours'] for d in last_month_data]), 2)
        mc['hours_diff'] = round(mc['this_avg_hours'] - mc['last_avg_hours'], 2)
        
        this_scores = [d['score'] for d in this_month_data if 'score' in d]
        last_scores = [d['score'] for d in last_month_data if 'score' in d]
        if this_scores and last_scores:
            mc['this_avg_score'] = round(avg(this_scores), 1)
            mc['last_avg_score'] = round(avg(last_scores), 1)
            mc['score_diff'] = round(mc['this_avg_score'] - mc['last_avg_score'], 1)
        
        this_deep = [d['deep'] for d in this_month_data if 'deep' in d]
        last_deep = [d['deep'] for d in last_month_data if 'deep' in d]
        if this_deep and last_deep:
            mc['this_avg_deep'] = round(avg(this_deep), 2)
            mc['last_avg_deep'] = round(avg(last_deep), 2)

    # Improvements
    improvements = []
    recent_30 = [d for d in data if d.get('hours') and d['date'] >= (today - timedelta(days=30)).strftime('%Y-%m-%d')]
    
    if recent_30:
        avg_hours = avg([d['hours'] for d in recent_30])
        if avg_hours < 7:
            improvements.append(f"直近30日の平均睡眠は{avg_hours:.1f}hで、推奨の7h未満です")
        
        # Check bedtime patterns by day of week
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        for dow in range(7):
            dow_data = [d for d in recent_30 if d.get('bedtime') and datetime.strptime(d['date'], '%Y-%m-%d').weekday() == dow]
            if len(dow_data) >= 2:
                def to_decimal(t):
                    h, m = map(int, t.split(':'))
                    return (h + 24 if h < 12 else h) + m/60
                avg_bed = avg([to_decimal(d['bedtime']) for d in dow_data])
                if avg_bed > 23.5:
                    improvements.append(f"{day_names[dow]}曜日の平均就寝が{int(avg_bed)}:{int((avg_bed%1)*60):02d}と遅い傾向")
        
        # Deep sleep trend
        recent_deep = [d for d in recent_30 if 'deep' in d]
        if len(recent_deep) >= 14:
            first_half = recent_deep[:len(recent_deep)//2]
            second_half = recent_deep[len(recent_deep)//2:]
            first_avg = avg([d['deep'] for d in first_half])
            second_avg = avg([d['deep'] for d in second_half])
            if second_avg < first_avg * 0.85:
                improvements.append(f"深い睡眠が減少傾向（{first_avg:.1f}h→{second_avg:.1f}h）")
            elif second_avg > first_avg * 1.15:
                improvements.append(f"深い睡眠が改善傾向（{first_avg:.1f}h→{second_avg:.1f}h）✓")
        
        # Score consistency
        scores = [d['score'] for d in recent_30 if 'score' in d]
        if scores:
            low_days = sum(1 for s in scores if s < 85)
            if low_days >= 5:
                improvements.append(f"直近30日でスコア85未満が{low_days}日あり")

    report['improvements'] = improvements

    # Streaks
    sleep_data = sorted([d for d in data if d.get('hours')], key=lambda d: d['date'], reverse=True)
    streak_7h = 0
    for d in sleep_data:
        if d['hours'] >= 7:
            streak_7h += 1
        else:
            break
    report['streaks']['days_7h_plus'] = streak_7h

    return report


def generate_obsidian_report(data: list[dict], report: dict) -> str:
    """Obsidianマークダウンレポートを生成"""
    today = datetime.now()
    week_num = today.isocalendar()[1]
    
    lines = [
        f"---",
        f"date: {today.strftime('%Y-%m-%d')}",
        f"type: sleep-report",
        f"tags: [sleep, report, weekly]",
        f"---",
        f"",
        f"# 🌙 睡眠レポート {today.strftime('%Y年%m月%d日')}",
        f"",
    ]

    # Weekly summary
    wk = report.get('weekly', {})
    if wk:
        lines.append("## 📊 今週の傾向")
        lines.append(f"- **平均睡眠時間**: {wk.get('avg_hours', '—')}h")
        if wk.get('avg_score'):
            lines.append(f"- **平均スコア**: {wk['avg_score']}")
        if wk.get('vs_last_week') is not None:
            diff = wk['vs_last_week']
            arrow = "↑" if diff > 0 else "↓" if diff < 0 else "→"
            lines.append(f"- **先週比**: {arrow} {abs(diff):.1f}h")
        if wk.get('best_day'):
            lines.append(f"- **ベスト**: {wk['best_day']}")
        lines.append("")

    # Monthly comparison
    mc = report.get('monthly_comparison', {})
    if mc:
        lines.append("## 📅 先月との比較")
        lines.append(f"| 項目 | {mc.get('last_month', '')} | {mc.get('this_month', '')} | 差分 |")
        lines.append("|---|---|---|---|")
        lines.append(f"| 平均睡眠 | {mc.get('last_avg_hours', '—')}h | {mc.get('this_avg_hours', '—')}h | {mc.get('hours_diff', 0):+.2f}h |")
        if mc.get('this_avg_score'):
            lines.append(f"| 平均スコア | {mc.get('last_avg_score', '—')} | {mc.get('this_avg_score', '—')} | {mc.get('score_diff', 0):+.1f} |")
        if mc.get('this_avg_deep'):
            lines.append(f"| 深い睡眠 | {mc.get('last_avg_deep', '—')}h | {mc.get('this_avg_deep', '—')}h | — |")
        lines.append("")

    # Improvements
    if report.get('improvements'):
        lines.append("## 💡 改善点・気づき")
        for imp in report['improvements']:
            lines.append(f"- {imp}")
        lines.append("")

    # Streaks
    streaks = report.get('streaks', {})
    if streaks.get('days_7h_plus', 0) > 0:
        lines.append(f"## 🔥 連続記録")
        lines.append(f"- 7時間以上の連続日数: **{streaks['days_7h_plus']}日**")
        lines.append("")

    lines.append(f"---")
    lines.append(f"*自動生成: {report.get('generated', '')}*")
    
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="総合ライフダッシュボード")
    parser.add_argument("--deploy", action="store_true", help="GitHub Pagesにデプロイ")
    args = parser.parse_args()

    print("📖 日記ファイルを読み込み中...")
    data = extract_all_data()
    
    has_sleep = [d for d in data if d.get('hours')]
    has_exercise = [d for d in data if d.get('exercise')]
    has_books = [d for d in data if d.get('books')]
    has_steps = [d for d in data if d.get('steps')]

    print(f"   → {len(data)} 日分のデータを抽出")
    print(f"      睡眠: {len(has_sleep)}日 / 筋トレ: {len(has_exercise)}日 / 歩数: {len(has_steps)}日 / 読書: {len(has_books)}日")

    # Stats
    if has_sleep:
        hours = [d['hours'] for d in has_sleep]
        scores = [d['score'] for d in has_sleep if 'score' in d]
        print(f"\n   📊 睡眠統計:")
        print(f"      平均: {sum(hours)/len(hours):.1f}h / 最長: {max(hours):.1f}h / 最短: {min(hours):.1f}h")
        if scores:
            print(f"      平均スコア: {sum(scores)/len(scores):.0f}")

    if has_books:
        finished = sum(1 for d in data for b in d.get('books', []) if b.get('finished'))
        print(f"   📚 読了: {finished}冊")

    # Sleep analysis report
    print("\n🧠 睡眠分析レポート生成中...")
    report = generate_sleep_report(data)
    
    if report.get('improvements'):
        print("   💡 改善点:")
        for imp in report['improvements']:
            print(f"      - {imp}")

    # Generate Obsidian report
    obsidian_md = generate_obsidian_report(data, report)
    report_path = VAULT_DIR / f"睡眠レポート_{datetime.now().strftime('%Y-%m-%d')}.md"
    report_path.write_text(obsidian_md, encoding='utf-8')
    print(f"\n📝 Obsidianレポート: {report_path}")

    # Reading summary
    reading_summary = build_reading_summary(data)
    print(f"   📚 ジャンル別: {', '.join(f'{g}:{c}' for g,c in sorted(reading_summary['genre_counts'].items(), key=lambda x:-x[1]))}")
    print(f"   📖 平均読了ペース: {reading_summary['avg_pace']}日/冊")

    # Generate HTML dashboard
    print("\n🎨 HTMLダッシュボード生成中...")
    data_json = json.dumps(data, ensure_ascii=False)
    report_json = json.dumps(report, ensure_ascii=False)
    reading_json = json.dumps(reading_summary, ensure_ascii=False)
    
    DOCS_DIR.mkdir(exist_ok=True)
    
    # Read HTML template
    template_path = SCRIPT_DIR / "dashboard_template.html"
    if template_path.exists():
        html = template_path.read_text(encoding='utf-8')
        html = html.replace('__DATA_JSON__', data_json)
        html = html.replace('__REPORT_JSON__', report_json)
        html = html.replace('__READING_JSON__', reading_json)
    else:
        print(f"   ⚠️ テンプレートが見つかりません: {template_path}")
        return

    index_path = DOCS_DIR / "index.html"
    index_path.write_text(html, encoding='utf-8')
    print(f"   ✓ {index_path}")

    # Also save to Vault for local viewing
    vault_html = VAULT_DIR / "睡眠ダッシュボード.html"
    vault_html.write_text(html, encoding='utf-8')
    print(f"   ✓ {vault_html}")

    # Generate Sleep App
    print("\n🌙 睡眠記録アプリ生成中...")
    sleep_template_path = SCRIPT_DIR / "sleep_template.html"
    if sleep_template_path.exists():
        sleep_data = [d for d in data if d.get('hours') or d.get('score')]
        sleep_json = json.dumps(sleep_data, ensure_ascii=False)
        sleep_html = sleep_template_path.read_text(encoding='utf-8')
        sleep_html = sleep_html.replace('__SLEEP_JSON__', sleep_json)
        sleep_path = DOCS_DIR / "sleep.html"
        sleep_path.write_text(sleep_html, encoding='utf-8')
        print(f"   ✓ {sleep_path}")
    else:
        print(f"   ⚠️ 睡眠テンプレートが見つかりません: {sleep_template_path}")

    # Deploy
    if args.deploy:
        print("\n🚀 GitHub Pagesにデプロイ中...")
        try:
            subprocess.run(["git", "add", "."], cwd=str(SCRIPT_DIR), check=True)
            subprocess.run(["git", "commit", "-m", f"update {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
                         cwd=str(SCRIPT_DIR), check=True)
            subprocess.run(["git", "push"], cwd=str(SCRIPT_DIR), check=True)
            print("   ✓ デプロイ完了！")
        except Exception as e:
            print(f"   ⚠️ デプロイ失敗: {e}")
    
    print("\n✅ 完了！")


if __name__ == "__main__":
    main()
